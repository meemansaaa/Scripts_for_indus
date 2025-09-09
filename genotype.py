import openpyxl as px
#change the column numbers accordingly

# col 1 = SNP/rsid
# col 2 = effect allele
# col 3 = alt allele
# col 4 = outcome (risk)
# col 5 = OR value
# col 6 = Genotype

def genotype(file):  #fn do create genotype from ref/alt alleles
    slash = '/'
    for i in range(2, rows+1, 3):
        eff1 = sheet_obj.cell(row=i, column=2).value   #type of eff3 is str
        alt1 = sheet_obj.cell(row=i, column=3).value
        if(type(eff1)==str and type(alt1)==str):  #to checkif type is str
            if(len(eff1)==1 and len(alt1)==1):
                g1 = alt1 + alt1 #for SNPs/SNVs
            elif((len(eff1)>1 and len(alt1)==1) or (len(eff1)==1 or len(alt1)>1)):
                g1 = alt1 + slash + alt1   #for indels
            else:
                g1 = 'NA'
            
        else:  
            g1= ' '    #ifnot str, i.e. if empty (in this case)
            #print('not string')
        #print(eff1,alt1)
        #print(g1) #type is str
        #print(g1)
        c7_1= sheet_obj.cell(row = i, column = 6)
        c7_1.value = g1   #putting the value in the cell
    #for i in range(3, rows+1, 3):
        eff2 = sheet_obj.cell(row=i+1, column=2).value
        alt2 = sheet_obj.cell(row=i+1, column=3).value
        if(type(eff2)==str and type(alt2)==str):
            if(len(eff2)==1 and len(alt2)==1):
                g2 = eff2  + alt2 
            elif((len(eff2)>1 and len(alt2)==1) or (len(eff2)==1 or len(alt2)>1)):
                g2 = eff2 + slash + alt2   
            else:
                g2 = 'NA'
        else:
            g2= ' ' 
        #print(g2)
        c7_2= sheet_obj.cell(row = i+1, column = 6)
        c7_2.value = g2
    #for i in range(4, rows+1, 3):
        eff3 = sheet_obj.cell(row=i+2, column=2).value
        alt3 = sheet_obj.cell(row=i+2, column=3).value
        if(type(eff2)==str and type(alt2)==str):
            if(len(eff3)==1 and len(alt3)==1):  
                g3 = eff3  + eff3    
            elif((len(eff3)>1 and len(alt3)==1) or (len(eff3)==1 or len(alt3)>1)):  
                g3 = eff3 + slash + eff3 
            else:
                g3 = 'NA'
        
        else:
            g3= ' '
        #print(g3)
        c7_3 = sheet_obj.cell(row = i+2, column = 6)
        c7_3.value = g3
        
    
def typical_or(file): #fn to make every the OR value of third entry as 1
    typical_val = 1
    for i in range(2,rows+1,3):
        or_col = sheet_obj.cell(row = i+2, column = 5)
        or_col.value = typical_val
        #testing:
        #print(sheet_obj.cell(row = i, column = 6).value)  
        #print(sheet_obj.cell(row = i+1, column = 6).value) 
        #print(sheet_obj.cell(row = i+2, column = 6).value)   


if __name__ == '__main__':
    file = 'no_genotyped_file.xlsx'
    wb_obj = px.load_workbook(file)
    sheet_obj = wb_obj.active # working in the active sheet

    # maximum rows and columns values
    rows = sheet_obj.max_row
    columns = sheet_obj.max_column
    print('Total rows:', rows)
    print('Total columns:', columns)

#testing:
    ff3 = sheet_obj.cell(row=2, column=2).value
    '''if(type(ff3)==str):
        print('string')
    else:
        print('not')'''
    
    #print(len(ff3))
    #print(ff3[1])

    genotype(file)
    #typical_or(file)
    wb_obj.save('genotyped_file3.xlsx') #pipe this file into typical_or
    



    



