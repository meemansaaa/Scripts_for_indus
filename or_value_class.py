import pandas as pd
import openpyxl as px # to read and write excel files

#change the column numbers accordingly
# col 1 = SNP/rsid
# col 2 = effect allele
# col 3 = alt allele
# col 4 = outcome (risk)
# col 5 = OR value
# col 6 = Genotype

#c2 = sheet_obj.cell(row = 2, column = 2)
# looping through each row in one column, and then writing out the classfied value in the next column

def or_class(file):
    for i in range(2,rows + 1):
        cell_obj = sheet_obj.cell(row=i, column=5) 
        #print(cell_obj.value)
        value = cell_obj.value  #saving the value from this cell to the variable
    #c2 = sheet_obj.cell(row = 2, column = 2)
    #print(value)
    #value2 = float(value)
    #type(value2)
    #print(value2)

    #classifying, while converting 'string' to 'float'
        if(type(value)==float or type(value)==int):
            if((float(value)) >= 1.09):
                category = 'Increased risk'
            elif((float(value)) < 1.0):
                category = 'Decreased risk'
            elif((float(value))>=1.0 and (float(value)<1.1)):
                category = 'Typical risk'
            else:
                category = 'NA'
        else:
            category = ' '
        c2 = sheet_obj.cell(row = i, column = 4)
        c2.value = category
       
    #iterating the column 2 through each row and saving it in c2
    # while assigning the category value to each row in the column
    

def typical_or(file): #fn to make every the OR value of third entry as 1
    typical_val = 1.0
    for i in range(2,rows+1,3):
        or_col = sheet_obj.cell(row = i+2, column = 5)
        or_col.value = typical_val
#saving and writing out the file   


if __name__=='__main__' :
    file = 'typical_dnawise3.xlsx'
    wb_obj = px.load_workbook(file)
    sheet_obj = wb_obj.active # working in the active sheet

# maximum rows and columns values
    rows = sheet_obj.max_row
    columns = sheet_obj.max_column
    print('Total rows:', rows)
    print('Total columns:', columns)



    #can run seperately or together
    #typical_or(file) #need to run this first
    or_class(file) # only then this
    
    wb_obj.save('classified2_dnawise3.xlsx')




