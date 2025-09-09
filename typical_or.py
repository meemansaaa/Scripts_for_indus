import openpyxl as px

def typical_or(file): #fn to make every the OR value of third entry as 1
    typical_val = 1.0
    for i in range(2,rows+1,3):
        or_col = sheet_obj.cell(row = i+2, column = 5)
        or_col.value = typical_val


if __name__=='__main__' :
    file = 'genotyped_dnawise3.xlsx'
    wb_obj = px.load_workbook(file)
    sheet_obj = wb_obj.active # working in the active sheet

# maximum rows and columns values
    rows = sheet_obj.max_row
    columns = sheet_obj.max_column
    print('Total rows:', rows)
    print('Total columns:', columns)

    typical_or(file) #need to run this first
    
    wb_obj.save('typical_dnawise3.xlsx') #pipe this file into or_value_class